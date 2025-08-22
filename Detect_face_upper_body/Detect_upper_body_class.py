from ultralytics import YOLO
from tqdm import tqdm
from rich.progress import Progress, BarColumn, MofNCompleteColumn, TimeRemainingColumn, TextColumn, TransferSpeedColumn, ProgressColumn
from rich.text import Text
from openpyxl import load_workbook

import os
import cv2
import pandas as pd
import numpy as np
import shutil

# Custom TransferSpeedColumn's unit
class FrameSpeedColumn(ProgressColumn):
    def render(self, task):
        speed = task.speed
        if speed is None:
            return Text(f"0 f/s", style="red")
        return Text(f"{speed:.2f} frames/s", style="red")

class PersonDetect:
    # Load model
    def __init__(self, model_name):
        self.model_name = model_name
        self.model = None
        self.export_data = []
        self.base_dir = None

    def load_model(self):
        self.model = YOLO(self.model_name)

    def image_label(self, image, text):
        font = cv2.FONT_HERSHEY_SIMPLEX
        scale = 0.5
        thickness = 1
        color = (225, 225, 225)
        bg_color = (0, 0, 0)

        h, w = image.shape[:2]
        org = (0, h - 2)

        (tw, th), _ = cv2.getTextSize(text, font, scale, thickness)
        cv2.rectangle(image, (org[0], org[1]-th), (org[0]+tw, org[1]), bg_color, -1)

        return cv2.putText(image, text, org, font, scale, color, thickness, cv2.LINE_AA)

    # Detect upper body
    def detect(self, video_path):

        if self.model is None:
            raise ValueError("Model is not loaded!")
        
        # Save direction for images of each video
        video_name = os.path.splitext(os.path.basename(video_path))[0]
        self.base_dir = os.path.join("Final_detect_report")
        data_dir = os.path.join(self.base_dir, "Data")
        os.makedirs(data_dir, exist_ok=True)

        data_test_dir = os.path.join(data_dir, "Data_test", "ShortVideo")
        frame_dir = os.path.join(data_dir, "Frames", video_name)
        detect_dir = os.path.join(data_dir, "Detect_output", video_name)
        merge_dir = os.path.join(data_dir, "Merged_output", video_name)
        os.makedirs(data_test_dir, exist_ok=True)
        os.makedirs(frame_dir, exist_ok=True)
        os.makedirs(detect_dir, exist_ok=True)
        os.makedirs(merge_dir, exist_ok=True)
        
        video_copy_path = os.path.join(data_test_dir, os.path.basename(video_path))
        if not os.path.exists(video_copy_path):
            shutil.copy2(video_path, video_copy_path)

        cap = cv2.VideoCapture(video_path)
        fps = int(cap.get(cv2.CAP_PROP_FPS))
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        real_total_frames = total_frames // int(fps*1)

        frame_id = 0
        saved = 0
        frame_list = []
        # with tqdm(total=total_frames, 
        #     desc=f"{video_name}", 
        #     unit=" frames", 
        #     ncols=70,
        #     bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{rate_fmt}, {remaining}]"
        # ) as pbar:

        with Progress(
            TextColumn("[progress.percentage]{task.percentage:>3.0f}%"), 
            BarColumn(),
            MofNCompleteColumn(),
            "[blue]frames",
            FrameSpeedColumn(),
            "etc:",                     
            TimeRemainingColumn(),
        ) as progress:
            task = progress.add_task("Processing", total=real_total_frames)
            while True:
                # Read frame from video
                ret, frame = cap.read()
                if not ret:
                    break
                frame_id += 1
                # Use YOLO model to detect for each frame

                if frame_id % int(fps*1) != 0:
                    continue                    
                results = self.model(frame, verbose=False)

                frame_path = os.path.join(frame_dir, f"img_{frame_id:04d}.jpg")
                frame_path_rel = os.path.relpath(frame_path, self.base_dir)
                cv2.imwrite(frame_path, frame)

                #Export to image with boudingbox's upper body
                for box in results[0].boxes: # YOLO detect a frame for each times
                    cls_id = int(box.cls[0]) # Take class id of object which is detected by YOLO
                    if cls_id == 0:          # YOLO use COCO dataset, class 0 = "person"

                        # Cut video into full body image
                        x1, y1, x2, y2 = map(int, box.xyxy[0])
                        w = x2 - x1 # Height of image
                        h = y2 - y1 # Width of image
                        cropped = frame[y1:y2, x1:x2]


                        if cropped.size > 0:
                            
                            # Draw boudingbox for upper body in each image
                            upper_body = int(h * 0.6)  # 60% of boudingbox
                            boxed = cropped.copy()
                            # cv2.rectangle(cropped, (0, 0), (w, upper_body), (255, 0, 0), 2)

                            # Save export image after cut and set boudingbox
                            save_path = os.path.join(detect_dir, f"img_{saved:04d}.jpg")
                            cv2.imwrite(save_path, cropped)
                            
                            merge_h = 300
                            frame_resized   = cv2.resize(frame,   (int(frame.shape[1] * merge_h / frame.shape[0]), merge_h))
                            cropped_resized = cv2.resize(cropped, (int(cropped.shape[1] * merge_h / cropped.shape[0]), merge_h))
                            boxed_resized   = cv2.resize(boxed,   (int(boxed.shape[1] * merge_h / boxed.shape[0]), merge_h))

                            frame_resized   = self.image_label(frame_resized,   "Frame")
                            cropped_resized = self.image_label(cropped_resized, "BB")
                            boxed_resized   = self.image_label(boxed_resized,   "Full Body")

                            merged = np.hstack((frame_resized, boxed_resized))
                            merged_path = os.path.join(merge_dir, f"img_{saved:04d}.jpg")
                            merged_path_rel = os.path.relpath(merged_path, self.base_dir)
                            cv2.imwrite(merged_path, merged)
                            saved += 1

                            time_frame = cap.get(cv2.CAP_PROP_POS_MSEC) / 1000.0
                            if frame_id not in frame_list:
                                frame_list.append(frame_id)
                                self.export_data.append({
                                    "NO": len(self.export_data)+1,
                                    "DATA INPUT": frame_path_rel,
                                    "INPUT'S DATA": f"{time_frame:.2f} s (frame: {frame_id} | {video_name})",
                                    # "OUTPUT DATA": merged_path_rel
                                })
                            else:
                                continue

                progress.update(task, advance=1)
                # pbar.update(1)
        print(f"{frame_list}")
        cap.release()
        print(f"---- Successful saved: {saved} -> {detect_dir} ----\n")

    def export_excel(self, video_path):

        self.base_dir = os.path.join("Final_detect_report")
        excel_path = os.path.join(self.base_dir, "detect_report.xlsx")

        if not self.export_data:
            print("Don't exist any data to export!!!")
            return

        df = pd.DataFrame(self.export_data)

        df.insert(0, "", "")
        df.to_excel(excel_path, index=False, startrow=1)

        wb = load_workbook(excel_path)
        ws = wb.active

        col_data_input = 3
        col_output = 5

        for row in range(3, len(df) + 3):
            # Hyperlink for DATA INPUT
            cell_input = ws.cell(row=row, column=col_data_input)
            file_input = cell_input.value
            if file_input:
                cell_input.hyperlink = file_input
                cell_input.style = "Hyperlink"

            # # Hyperlink for OUTPUT DATA
            # cell_output = ws.cell(row=row, column=col_output)
            # file_output = cell_output.value
            # if file_output:
            #     cell_output.hyperlink = file_output
            #     cell_output.style = "Hyperlink"

        wb.save(excel_path)
        print(f"Export data successful with hyperlinks: {excel_path}\n")
